#!/usr/bin/env python3
# Hayley Wragg 2019-03-20
''' The code saves the values for the parameters in a ray tracer '''
import numpy as np
import math as ma
import sys
import os
import pickle
import openpyxl as wb

def DeclareParameters(SN):
  '''All input parameters for the ray-launching method are entered in
  this function which will then save them inside a Parameters folder.

  * deltheta- The array of angle spacings.

  * Nre - Number of reflections
  * Ns - Number of steps to split longest axis.
  * l1 - Interior obstacle scale
  * l2 - Boundary scale.
  * triangle1 - First interior obstacle
  * ...
  * triangleN - Last interior obstacle
  * OuterBoundary1 - First obstacle forming the boundary of the \
  environment
  * ...
  * OuterBoundaryN - Last obstacle forming the boundary of the \
  environment.

  :return: 0 if successfully completed.

  '''
  # -------------------------------------------------------------------
  # INPUT PARAMETERS FOR RAY LAUNCHER----------------------------------
  # -------------------------------------------------------------------

  #print('Saving ray-launcher parameters')
  InBook     =wb.load_workbook(filename=SN,data_only=True)
  AngSpacestr='AngleSpacing'
  SimParstr  ='SimulationParameters'
  Direcstr   ='Directions'
  Obststr    ='Obstacles'
  OutBstr    ='OuterBoundary'
  NTriObstr  ='NTriObst'
  NTriOutstr ='NTriOut'
  Angspace   =InBook[AngSpacestr]
  SimPar     =InBook[SimParstr]
  Direc      =InBook[Direcstr]
  Obst       =InBook[Obststr]
  OutB       =InBook[OutBstr]
  NTriObSh   =InBook[NTriObstr]
  NTriOutSh  =InBook[NTriOutstr]

  testnum    =SimPar.cell(row=17,column=3).value
  roomnumstat=SimPar.cell(row=18,column=3).value
  timetest   =SimPar.cell(row=19,column=3).value

  deltheta=np.array([])
  nrays=Angspace.max_row-1
  nrays=1
  SimPar.cell(row=12,column=3).value=nrays


  for j in range(nrays):
    deltheta=np.append(deltheta,Angspace.cell(row=j+2,column=1).value)
    #np.pi*np.array([1/3])#,1/5,1/7,1/8,1/9,1/12,1/14,1/16,1/18,1/19,1/20,1/22,1/25,1/36])
  Nra=np.ones(nrays,dtype=int)
  Nre=int(SimPar.cell(row=2,column=3).value )           # Number of reflections
  Ns=int(SimPar.cell(row=3,column=3).value )             # Number of steps on longest axis.
  #for i,j in product(range(Nre*Ns),range(Nre*Ns)):
  #   option=np.append(option,i*straight+j*diagnoal)
  split=SimPar.cell(row=4,column=3).value           # Number of steps through each mesh square
  l1=SimPar.cell(row=5,column=3).value            # Interior obstacle scale
  l2=SimPar.cell(row=6,column=3).value            # Outer Boundary length scale
  InnerOb=SimPar.cell(row=7,column=3).value         # Indicator of whether the inner obstacles should be used
  NtriOut=np.array([])# This will be the number of triangles forming each plane surface in the outer boundary
  NtriOb=np.array([]) # This will be the number of triangles forming each plane surface in the obstacle list

  Nbox=Obst.max_row-1
  Oblist=np.array([])
  for j in range(Nbox):
    ## Obstacles are all triangles in 3D.
    Box=Obst.cell(row=j+2,column=3).value
    Tri=Obst.cell(row=j+2,column=4).value
    if Box:
        xmi=Obst.cell(row=j+2,column=5).value
        xma=Obst.cell(row=j+2,column=6).value
        ymi=Obst.cell(row=j+2,column=7).value
        yma=Obst.cell(row=j+2,column=8).value
        zmi=Obst.cell(row=j+2,column=9).value
        zma=Obst.cell(row=j+2,column=10).value
        if j==0:
          Oblist=BoxBuild(xmi,xma,ymi,yma,zmi,zma)
        else:
          Oblist=np.vstack((Oblist.astype(float),BoxBuild(xmi,xma,ymi,yma,zmi,zma).astype(float)))
        # In a box all surfaces are formed of two triangles
        NTribox=2*np.ones(6)
        NtriOb=np.append(NtriOb,NTribox)
    elif Tri:
        p0=np.ones(3)
        p1=np.ones(3)
        p2=np.ones(3)
        p0[0]=Obst.cell(row=j+2,column=11).value
        p0[1]=Obst.cell(row=j+2,column=12).value
        p0[2]=Obst.cell(row=j+2,column=13).value
        p1[0]=Obst.cell(row=j+2,column=14).value
        p2[1]=Obst.cell(row=j+2,column=15).value
        p1[2]=Obst.cell(row=j+2,column=16).value
        p2[0]=Obst.cell(row=j+2,column=17).value
        p2[1]=Obst.cell(row=j+2,column=18).value
        p2[2]=Obst.cell(row=j+2,column=19).value
        Tri=np.array([p0,p1,p2]).astype(float)
        if j==1:
          Oblist=Tri
        else:
          Oblist=np.vstack((Oblist.astype(float),Tri))
        # In a box all surfaces are formed of two triangles
        NtriOb=np.append(NtriOb,1)
  for j in range(len(NtriOb)):
    NTriObSh.cell(row=j+1,column=1).value=NtriOb[j]

  #- Outer Boundary -
  # 3D co-ordinates forming a closed boundary.
  Nbox=OutB.max_row-1
  OuterBoundary=np.array([])
  for j in range(Nbox):
    ## Obstacles are all triangles in 3D.
    Box=OutB.cell(row=j+2,column=3).value
    Tri=OutB.cell(row=j+2,column=4).value
    if Box:
        xmi=OutB.cell(row=j+2,column=5).value
        xma=OutB.cell(row=j+2,column=6).value
        ymi=OutB.cell(row=j+2,column=7).value
        yma=OutB.cell(row=j+2,column=8).value
        zmi=OutB.cell(row=j+2,column=9).value
        zma=OutB.cell(row=j+2,column=10).value
        Bound=BoxBuild(xmi,xma,ymi,yma,zmi,zma)
        if j==0:
          OuterBoundary=Bound
        else:
          OuterBoundary=np.vstack((OuterBoundary,Bound))
        # In a box all surfaces are formed of two triangles
        NTribox=2*np.ones(6)
        NtriOut=np.append(NtriOut,NTribox)
    elif Tri:
        p0=np.ones(3)
        p1=np.ones(3)
        p2=np.ones(3)
        p0[0]=OutB.cell(row=j+2,column=11).value
        p0[1]=OutB.cell(row=j+2,column=12).value
        p0[2]=OutB.cell(row=j+2,column=13).value
        p1[0]=OutB.cell(row=j+2,column=14).value
        p2[1]=OutB.cell(row=j+2,column=15).value
        p1[2]=OutB.cell(row=j+2,column=16).value
        p2[0]=OutB.cell(row=j+2,column=17).value
        p2[1]=OutB.cell(row=j+2,column=18).value
        p2[2]=OutB.cell(row=j+2,column=19).value
        Tri=np.array([p0,p1,p2])
        if j==0:
          OuterBoundary=Tri
        else:
          OuterBoundary=np.vstack((OuterBoundary,Tri))
        # In a box all surfaces are formed of two triangles
        NtriOut=np.append(NtriOut,1)
  for j in range(len(NtriOut)):
    NTriOutSh.cell(row=j+1,column=1).value=NtriOut[j]

  # -Router location -co-ordinate of three real numbers
  Tx=np.array([SimPar['D8'].value,SimPar['E8'].value,SimPar['F8'].value])

  Nout =int(np.sum(NtriOut))
  Nobst=int(np.sum(NtriOb))
  Nob  =int(Nout+InnerOb*Nobst)
  ObCooStr="ObstacleCoords"
  try:
    ObstCoor=InBook[ObCooStr]
  except KeyError:
    InBook.create_sheet(ObCooStr)
    ObstCoor=InBook[ObCooStr]
  for j in range(Nob):
    if j+1<Nout:
      Obstr='OuterBoundar%d'%j
      ObstCoor.cell(row=3*j+1,column=1).value=Obstr
      ObstCoor.cell(row=3*j+1,column=2).value=OuterBoundary[j,0,0]
      ObstCoor.cell(row=3*j+1,column=3).value=OuterBoundary[j,0,1]
      ObstCoor.cell(row=3*j+1,column=4).value=OuterBoundary[j,0,2]
      ObstCoor.cell(row=3*j+2,column=2).value=OuterBoundary[j,1,0]
      ObstCoor.cell(row=3*j+2,column=3).value=OuterBoundary[j,1,1]
      ObstCoor.cell(row=3*j+2,column=4).value=OuterBoundary[j,1,2]
      ObstCoor.cell(row=3*j+3,column=2).value=OuterBoundary[j,2,0]
      ObstCoor.cell(row=3*j+3,column=3).value=OuterBoundary[j,2,1]
      ObstCoor.cell(row=3*j+3,column=4).value=OuterBoundary[j,2,2]
    if Nout<=j+1<Nout+Nobst:
      Obstr='Obstacle%d'%j
      ObstCoor.cell(row=3*j+1,column=1).value=Obstr
      ObstCoor.cell(row=3*j+1,column=2).value=Oblist[j-Nout,0,0]
      ObstCoor.cell(row=3*j+1,column=3).value=Oblist[j-Nout,0,1]
      ObstCoor.cell(row=3*j+1,column=4).value=Oblist[j-Nout,0,2]
      ObstCoor.cell(row=3*j+2,column=2).value=Oblist[j-Nout,1,0]
      ObstCoor.cell(row=3*j+2,column=3).value=Oblist[j-Nout,1,1]
      ObstCoor.cell(row=3*j+2,column=4).value=Oblist[j-Nout,1,2]
      ObstCoor.cell(row=3*j+3,column=2).value=Oblist[j-Nout,2,0]
      ObstCoor.cell(row=3*j+3,column=3).value=Oblist[j-Nout,2,1]
      ObstCoor.cell(row=3*j+3,column=4).value=Oblist[j-Nout,2,2]
  runplottype=SimPar.cell(row=15,column=3).value
  Heatmappattern=SimPar.cell(row=16,column=3).value

  LOS=SimPar.cell(row=10,column=3).value    # LOS=1 for LOS propagation, LOS=0 for reflected propagation
  PerfRef=SimPar.cell(row=11,column=3).value # Perfect reflection has no loss and ignores angles.

  # -------------------------------------------------------------------
  # CALCULATED PARAMETERS TO SAVE
  # -------------------------------------------------------------------
  # CALCULATE RELATIVE MESHWIDTH
  roomlengthscale=l2*abs(np.amax(OuterBoundary)-np.amin(OuterBoundary)) # SCALE WITHIN THE UNIT CO-ORDINATES.
  SimPar.cell(row=6,column=3).value=roomlengthscale
  #OuterBoundary=OuterBoundary/roomlengthscale
  #Oblist=Oblist/roomlengthscale
  h=1.0/Ns
  SimPar.cell(row=14,column=3).value=h

  if not os.path.exists('./Parameters'):
    os.makedirs('./Parameters')
  # CALCULATE ANGLE SPACING
  for j in range(0,nrays):
    xysteps       =int(ma.ceil(abs(2.0*np.pi/deltheta[j])))
    theta1        =np.linspace(0.0,2*np.pi,num=int(xysteps), endpoint=False) # Create an array of all the angles
    deltheta[j]   =theta1[1]-theta1[0]
    zsteps        =int(ma.ceil(abs(np.pi/(deltheta[j]))))
    Nra[j]=2
    Nraout=np.array([])
    start=0
    for k in range(-int(zsteps/2),int(zsteps/2)+1):
      mid=(np.cos(deltheta[j])-np.sin(k*deltheta[j])**2)/(np.cos(deltheta[j]*k)**2)
      if abs(mid)>1:
        pass
      else:
        bot=ma.acos(mid)
        xyk=int(2*np.pi/bot)
        if xyk<=1:
          break
        Nra[j]+=xyk
        theta1        =np.linspace(0.0,2*np.pi,num=int(xyk), endpoint=False) # Create an array of all the angles
        co=np.cos(k*deltheta[j])
        si=np.sin(k*deltheta[j])
        updirecs     =np.c_[co*np.cos(theta1),co*np.sin(theta1),si*np.ones((xyk,1))]
        #downdirecs   =np.c_[co*np.cos(theta1),co*np.sin(theta1),-si*np.ones((xyk,1))]
        if start==0:
          coords=updirecs
          start=1
        else:
          #coords  =np.r_[coords,downdirecs]
          coords  =np.r_[updirecs,coords]
    if len(coords)<=1:
      Nraout=Nra[j+1:]
      pass
    else:
      Nraout=Nra
      directions=np.zeros((Nra[j],4))
      directions[1:-1]=np.c_[coords,np.zeros((Nra[j]-2,1))]
      directions[0] =np.array([0.0,0.0, 1.0,0.0])
      directions[-1]=np.array([0.0,0.0,-1.0,0.0])
      directionname='Parameters/Directions%d.npy'%j
      np.save(directionname,directions)
    Angspace.cell(row=j+2,column=2).value=Nra[j]
    DirecStr='Directions%d'%(Nra[j])
    try:
      DirecSh=InBook[DirecStr]
    except KeyError:
      InBook.create_sheet(DirecStr)
      DirecSh=InBook[DirecStr]
    for i in range(int(Nra[j])):
      DirecSh.cell(row=i+1,column=1).value=directions[i,0]
      DirecSh.cell(row=i+1,column=2).value=directions[i,1]
      DirecSh.cell(row=i+1,column=3).value=directions[i,2]


  # COMBINE THE RAY-LAUNCHER PARAMETERS INTO ONE ARRAY
  RTPar=np.array([Nre,h,roomlengthscale,split])

  print('Number of rays ', Nraout,'Number of reflections ', Nre,'Mesh spacing ', h)
  print('Angle spacing ', deltheta)
  #print('Origin of raytracer ', Tx)

  # --------------------------------------------------------------------
  # SAVE THE PARAMETERS IN A FOLDER TITLED `Parameters`
  # --------------------------------------------------------------------
  np.save('Parameters/testnum.npy',testnum)
  np.save('Parameters/roomnumstat.npy',roomnumstat)
  np.save('Parameters/timetest.npy',timetest)
  np.save('Parameters/Raytracing.npy',RTPar)
  np.save('Parameters/Nra.npy',Nraout)
  np.save('Parameters/delangle.npy',deltheta)
  if InnerOb:
    Oblist=np.concatenate((OuterBoundary,Oblist))
  else:
    Oblist=OuterBoundary
  np.save('Parameters/Obstacles.npy',Oblist)
  np.save('Parameters/NtriOb.npy',NtriOb)
  np.save('Parameters/InnerOb.npy',InnerOb)
  np.save('Parameters/OuterBoundary.npy',OuterBoundary)
  np.save('Parameters/NtriOut.npy',NtriOut)
  np.save('Parameters/Nob.npy',Nob)
  np.save('Parameters/Ns.npy',Ns)
  np.save('Parameters/Nob.npy',Nob)
  np.save('Parameters/Origin.npy',Tx)
  np.save('Parameters/LOS.npy',LOS)
  np.save('Parameters/PerfRef.npy',PerfRef)

  text_file = open('Parameters/runplottype.txt', 'w')
  n = text_file.write(runplottype)
  text_file.close()
  text_file = open('Parameters/Heatmapstyle.txt', 'w')
  n = text_file.write(Heatmappattern)
  text_file.close()
  #print('------------------------------------------------')
  #print('Geometrical parameters saved')
  #print('------------------------------------------------')
  InBook.save(filename=SN)
  return 0

def ObstacleCoefficients(SN,index=0):
  ''' Input the paramters for obstacles and the antenna. To ensure \
  arrays are of the right length for compatibility for the \
  ray-launcher retrieve the ray-launching parameters in \
  :py:func:`DeclareParameters()`

  Load:

  * 'Obstacles.npy'     -Co-ordinates of obstacles in the room
  * 'OuterBoundary.npy' - Co-ordinates of the walls of the room
  * 'Raytracing.npy'    -[Nra (number of rays), Nre (number of reflections), \
  h (relative meshwidth)]

  Calculate:

  * Nob=len([Obstacles,OuterBoundary])

  Input:

  * `Freespace` -[mu0 (permeability of air), \
  eps0 (permittivity of air),Z0 (characteristic impedance of air), \
  c (speed of light)]
  * `frequency` - :math:`\\omega` angular frequency of the wave out \
  the antenna.
  * `mur`       - :math:`\\mu_r` The relative permeability for all obstacles. \
  This should be an array with the same number of terms as the number \
  of obstacles Nob.
  * `epsr`     - :math:`\\epsilon_r` The relative permittivity for each obstacle. \
  This should be an array with the same number of terms as the number \
  of obstacles Nob.
  * `sigma`     - :math:`\\sigma` The electrical conductivity of the obstacles. \
  This should be an array with the same number of terms as the number \
  of obstacles.
  * `Gt`        - The gains of the antenna. The should be an array with \
  the same number of terms as the number of rays Nra.

  Calculate:

  * `eps0`   - :math:`\\epsilon_0=\\frac{1}{\\mu_0 c^2}`  permittivity of \
  freespace.
  * `Z0`     - :math:`Z_0=\\sqrt{\\frac{\\mu_0}{\\epsilon_0}}` characteristic \
  impedance of freespace.
  * `refindex` - The refreactive index \
  :math:`n=\\sqrt{\\mu_r\\epsilon_r}`
  * `Znobrat`- The relative impedance of the obstacles given by,
    :math:`\\hat{Z}_{Nob}=\\frac{Z_{Nob}}{Z_0}`. The impedance of each \
    obstacle :math:`Z_{Nob}` is given by \
    :math:`Z_{Nob}=\\sqrt{\\frac{i\\omega\\mu_0\\mu_r}{\\sigma+i\\epsilon_0\\epsilon_r}}`.


  The Znobrat and refindex terms are then reformatted so that they \
  repeat Nre times with an extra term. The extra term corresponds to \
  the line of sight path. This makes them the same length as a column \
  in a matrix in a :py:class:`DictionarySparseMatrix.DS`. \
  Each term corresponds to a possible obstacle reflection combination.

  The Gains matrix is also reformated to that it repeats (Nre+1) times. \
  This corresponds to every possible ray reflection number combination \
  This makes them the same length as a row in a matrix in a \
  :py:class:`DictionarySparseMatrix.DS`. \
  Each term corresponds to a possible obstacle reflection combination.

  Save:
  * `frequency.npy`- The angular frequency :math:`\\omega`.
  * `refindex.npy` - The refractive index of the obstacles.
  * `Znobrat.npy`  - The relative characteristic impedance.
  * `TxGains.npy`  - The gains of the antenna.
  * `Freespace.npy`- The freespace parameters.

  .. code::

     Freespace=np.array([mu0,eps0,Z0,c])

  :return: 0 if successfully completed.

  '''
  #print('Saving the physical parameters for obstacles and antenna')

  # -------------------------------------------------------------------
  # RETRIEVE RAY LAUNCHER PARAMETERS FOR ARRAY LENGTHS-----------------
  # -------------------------------------------------------------------
  InBook     =wb.load_workbook(filename=SN,data_only=True)
  AirMat     ='AirParameters'
  ObstMat    ='ObstacleMaterial'
  Simparstr  ='SimulationParameters'
  Obststr    ='Obstacles'
  OutBstr    ='OuterBoundary'
  Obst       =InBook[Obststr]
  OutB       =InBook[OutBstr]
  Air        =InBook[AirMat]
  ObstPar    =InBook[ObstMat]
  SimPar     =InBook[Simparstr]
  if not os.path.exists('./Parameters/'):
    os.makedirs('./Parameters/')
  RTPar         =np.load('Parameters/Raytracing.npy')
  Nre,h,L,split =RTPar
  Oblist        =np.load('Parameters/Obstacles.npy')
  Nra           =np.load('Parameters/Nra.npy')
  if isinstance(Nra, (float,int,np.int32,np.int64, np.complex128 )):
      nra=np.array([Nra])
  else:
      nra=len(Nra)
  Nre=int(RTPar[0])                           # Number of reflections
  Nob=np.load('Parameters/Nob.npy')           # The Number of obstacle.
  Nrs =SimPar.cell(row=13,column=3).value

  # -------------------------------------------------------------------
  # INPUT PARAMETERS FOR POWER CALCULATIONS----------------------------
  # -------------------------------------------------------------------

  # PHYSICAL CONSTANTS Do not change for obstacles for frequency
  mu0=complex(Air.cell(row=2,column=1).value) #4*np.pi*1E-6
  c  =float(Air.cell(row=2,column=3).value) #2.99792458E+8


  # ANTENNA PARAMETERS
  #-----------------------------------------------------------------------
  # Gains of the rays
  for j in range(0,nra):
    Gt=np.ones((Nra[j],1),dtype=np.complex128)
    gainname=str('Parameters/Tx'+str(Nra[j])+'Gains'+str(index)+'.npy')
    np.save(gainname, Gt)
  frequency     =float(Air.cell(row=2,column=4).value)              # 2.79 GHz #FIXME make this a table and choose a frequency option
  khat          =frequency*L/c                 # Non-dimensional wave number
  Air.cell(row=2,column=5).value=khat
  lam           =(2*np.pi)/khat                # Non-dimensional wavelength
  Air.cell(row=2,column=6).value=lam
  Antpar        =np.array([khat,lam,L])
  Pol           =np.array([Air.cell(row=2,column=7).value,Air.cell(row=3,column=7).value])
  # CALCULATE FREESPACE PARAMETERS

  eps0=1/(mu0*c**2)  #8.854187817E-12
  Air.cell(row=2,column=2).value=str(eps0)
  Z0=(mu0/eps0)**0.5 #120*np.pi Characteristic impedance of free space.
  Air.cell(row=2,column=8).value=str(Z0)

  # STORE FREESPACE PARAMETERS IN ONE ARRAY
  Freespace=np.array([mu0,eps0,Z0,c])
  # OBSTACLE CONTSTANTS
  #----------------------------------------------------------------------
  # Relative Constants for the obstacles
  NtriOut=np.load('Parameters/NtriOut.npy')
  NtriOb=np.load('Parameters/NtriOb.npy')

  InOb=SimPar.cell(row=7,column=3).value
  Nobst=OutB.max_row-1
  NOut=Obst.max_row-1
  Ns=Nobst*InOb+NOut
  mur   =np.array([complex(ObstPar.cell(row=2,column=2).value)])
  epsr  =np.array([complex(ObstPar.cell(row=2,column=3).value)])
  sigma =np.array([complex(ObstPar.cell(row=2,column=4).value)])
  top=frequency*mu0*mur[-1]*1j
  bottom=sigma[-1]+eps0*frequency*epsr[-1]*1j
  Znob =np.sqrt(top/bottom)                    # Wave impedance of the obstacles
  Znobrat=np.array([Znob/Z0])
  ObstPar.cell(row=j+2,column=5).value=str(Znob/Z0)
  refindex=np.array([mur[-1]*epsr[-1]])
  ObstPar.cell(row=2,column=6).value=str(refindex[-1])
  for j in range(0,Ns):
    if j+1<=NOut:
      for i in range(1,int(NtriOut[j]*(Obst.cell(row=j+2,column=3).value*6+Obst.cell(row=j+2,column=4).value))):
        mur   =np.append(mur,complex(ObstPar.cell(row=j+2,column=2).value))
        epsr  =np.append(epsr ,complex(ObstPar.cell(row=j+2,column=3).value))
        sigma =np.append(sigma,complex(ObstPar.cell(row=j+2,column=4).value))
        top=frequency*mu0*mur[-1]*1j
        bottom=sigma[-1]+eps0*frequency*epsr[-1]*1j
        Znob =np.sqrt(top/bottom)                    # Wave impedance of the obstacles
        Znobrat=np.append(Znobrat,Znob/Z0)
        ObstPar.cell(row=j+2,column=5).value=str(Znob/Z0)
        if i > Nrs*np.sum(NtriOut[:(Nrs-1)])+1:
          refindex=np.append(refindex,0+0j)
        else:
          refindex=np.append(refindex,mur[-1]*epsr[-1])
        ObstPar.cell(row=2,column=6).value=str(refindex[-1])
    if NOut+Nobst+1>j+1>NOut:
      for i in range(int(NtriOb[j]*(Obst.cell(row=j-NOut+2,column=3).value*6+Obst.cell(row=j-NOut+2,column=4).value))):
        mur   =np.append(mur  ,complex(ObstPar.cell(row=j+2,column=2).value))
        epsr  =np.append(epsr ,complex(ObstPar.cell(row=j+2,column=3).value))
        sigma =np.append(sigma,complex(ObstPar.cell(row=j+2,column=4).value))
        top    =frequency*mu0*mur[-1]*1j
        bottom =sigma[-1]+eps0*frequency*epsr[-1]*1j
        Znob   =np.sqrt(top/bottom)                    # Wave impedance of the obstacles
        Znobrat=np.append(Znobrat,Znob/Z0)
        ObstPar.cell(row=j+2,column=5).value=str(Znob/Z0)
        if i>Nrs*np.sum(NtriOut[:(Nrs-1)]):
          refindex=np.append(refindex,0+0j)
        else:
          refindex=np.append(refindex,mur[-1]*epsr[-1])
        ObstPar.cell(row=2,column=6).value=str(refindex[-1])
  # CALCULATE OBSTACLE PARAMETERS
  PerfRef=SimPar.cell(row=11,column=3).value
  if PerfRef:
    refindex=np.zeros(Nob,dtype=np.complex128)          # Perfect Relection
    refindex[0]=1 #np.sqrt(np.multiply(mur[0],epsr[0]))     # Refractive index of the obstacles
    refindex[1]=1
  # --------------------------------------------------------------------
  # SAVE THE PARAMETERS
  # --------------------------------------------------------------------
  np.save('Parameters/Freespace'+str(index)+'.npy',Freespace)
  np.save('Parameters/frequency'+str(index)+'.npy',frequency)
  np.save('Parameters/lam'+str(index)+'.npy',lam)
  np.save('Parameters/khat'+str(index)+'.npy',khat)
  np.save('Parameters/Antpar'+str(index)+'.npy',Antpar)
  np.save('Parameters/Znobrat'+str(index)+'.npy',Znobrat)
  np.save('Parameters/refindex'+str(index)+'.npy',refindex)
  np.save('Parameters/Pol'+str(index)+'.npy',Pol)
  #print('------------------------------------------------')
  #print('Material parameters saved')
  #print('------------------------------------------------')
  InBook.save(filename=SN)
  return 0

def BoxBuild(xmi,xma,ymi,yma,zmi,zma):
  ''' Input the inimum and maximum x,y, and z co-ordinates which will form a Box.
  :param xmi: The minimum x co-ordinate.
  :param xma: The maximum x co-ordinate.
  :param ymi:The minimum y co-ordinate.
  :param yma: The maximum y co-ordinate.
  :param zmi: The minimum z co-ordinate.
  :param zma: The maximum z co-ordinate.

  .. code::

       Box=[T0,T1,...T12]
       TJ=[p0J,p1J,p2J]
       p0J=[x0J,y0J,z0J]
       p1J=[x1J,y1J,z1J]
       p2J=[x2J,y2J,x2J]

  :rtype: 12 x 3 x 3 numpy array.
  :returns: Box
  '''
  # The faces in the y=ymi plane
  triangle1 =np.array([(xmi,ymi,zmi),(xma,ymi,zmi),(xmi,ymi,zma)])
  triangle2 =np.array([(xmi,ymi,zma),(xma,ymi,zma),(xma,ymi,zmi)])
  # The faces in the x=xmi plane
  triangle3 =np.array([(xmi,ymi,zmi),(xmi,yma,zmi),(xmi,ymi,zma)])
  triangle4 =np.array([(xmi,ymi,zma),(xmi,yma,zmi),(xmi,yma,zma)])
  # The faces in the z=zmi plane
  triangle5 =np.array([(xmi,ymi,zmi),(xma,yma,zmi),(xmi,yma,zmi)])
  triangle6 =np.array([(xmi,ymi,zmi),(xma,yma,zmi),(xma,ymi,zmi)])
  # The faces in the z=zma plane
  triangle7 =np.array([(xmi,ymi,zma),(xma,yma,zma),(xma,ymi,zma)])
  triangle8 =np.array([(xmi,ymi,zma),(xma,yma,zma),(xmi,yma,zma)])
  # The faces in the x=xma plane
  triangle9 =np.array([(xma,ymi,zma),(xma,yma,zma),(xma,yma,zmi)])
  triangle10=np.array([(xma,ymi,zma),(xma,ymi,zmi),(xma,yma,zmi)])
  # The faces in the y=yma plane
  triangle11=np.array([(xma,yma,zmi),(xma,yma,zma),(xmi,yma,zma)])
  triangle12=np.array([(xma,yma,zmi),(xmi,yma,zmi),(xmi,yma,zma)])
  # Put the triangular faces into an array called Box.
  Box=np.array([triangle1,triangle2,triangle3,triangle4,triangle5,
  triangle6,triangle7,triangle8,triangle9,triangle10,triangle11,
  triangle12])
  return Box

if __name__=='__main__':
  np.set_printoptions(precision=3)
  print('Running  on python version')
  print(sys.version)
  Sheetname='InputSheet.xlsx'
  out=DeclareParameters(Sheetname)
  out=ObstacleCoefficients(Sheetname)

  exit()



