#!/usr/bin/env python3
# Hayley Wragg 2019-03-20
''' The code saves the values for the parameters in a ray tracer '''
import numpy as np
import math as ma
import sys
import os
import pickle
import openpyxl as wb
import logging
import num2words as nw

epsilon=sys.float_info.epsilon

def DeclareParameters(SN):
  '''All input parameters for the ray-launching method are entered in
  this function which will then save them inside a Parameters folder.

  * deltheta- The array of angle spacings.

  * Nre - Number of reflections
  * Ns - Number of steps to split longest axis.
  * l1 - Interior obstacle scale
  * l2 - Boundary scale.
  * triangle1 - First interior obstacle
  * ...
  * triangleN - Last interior obstacle
  * OuterBoundary1 - First obstacle forming the boundary of the \
  environment
  * ...
  * OuterBoundaryN - Last obstacle forming the boundary of the \
  environment.

  :return: 0 if successfully completed.

  '''
  # -------------------------------------------------------------------
  # INPUT PARAMETERS FOR RAY LAUNCHER----------------------------------
  # -------------------------------------------------------------------

  #print('Saving ray-launcher parameters')
  print(SN)
  InBook     =wb.load_workbook(filename=SN,data_only=True)
  AngSpacestr='AngleSpacing'
  SimParstr  ='SimulationParameters'
  Obststr    ='Obstacles'
  OutBstr    ='OuterBoundary'
  NTriObstr  ='NTriObst'
  NTriOutstr ='NTriOut'
  Angspace   =InBook[AngSpacestr]
  SimPar     =InBook[SimParstr]
  Obst       =InBook[Obststr]
  OutB       =InBook[OutBstr]
  NTriObSh   =InBook[NTriObstr]
  NTriOutSh  =InBook[NTriOutstr]

  testnum    =SimPar.cell(row=17,column=3).value
  roomnumstat=SimPar.cell(row=18,column=3).value
  timetest   =SimPar.cell(row=19,column=3).value
  ResOn      =SimPar.cell(row=20,column=3).value
  MaxInter   =SimPar.cell(row=21,column=3).value
  logon      =SimPar.cell(row=25,column=3).value

  deltheta=np.array([])
  nrays=Angspace.max_row-1

  nrays=1
  SimPar.cell(row=12,column=3).value=nrays


  for j in range(nrays):
    deltheta=np.append(deltheta,Angspace.cell(row=j+2,column=1).value)
    #np.pi*np.array([1/3])#,1/5,1/7,1/8,1/9,1/12,1/14,1/16,1/18,1/19,1/20,1/22,1/25,1/36])
  Nra=np.ones(nrays,dtype=int)
  Nrs =SimPar.cell(row=13,column=3).value
  Nre=3
  int(SimPar.cell(row=2,column=3).value )=Nre     # Number of reflections
  Ns=int(SimPar.cell(row=3,column=3).value )      # Number of steps on longest axis.
  split=SimPar.cell(row=4,column=3).value         # Number of steps through each mesh square
  l1=SimPar.cell(row=5,column=3).value            # Interior obstacle scale
  l2=SimPar.cell(row=6,column=3).value            # Outer Boundary length scale
  InnerOb=0
  SimPar.cell(row=7,column=3).value=InnerOb       # Indicator of whether the inner obstacles should be used
  NtriOut=np.array([])# This will be the number of triangles forming each plane surface in the outer boundary
  NtriOb=np.array([]) # This will be the number of triangles forming each plane surface in the obstacle list

  Nbox=Obst.max_row-1
  Oblist=np.array([])
  for j in range(Nbox):
    ## Obstacles are all triangles in 3D.
    Box=Obst.cell(row=j+2,column=3).value
    Surf=Obst.cell(row=j+2,column=4).value
    Tri=Obst.cell(row=j+2,column=5).value
    if Box:
      xmi=Obst.cell(row=j+2,column=6).value
      xma=Obst.cell(row=j+2,column=7).value
      ymi=Obst.cell(row=j+2,column=8).value
      yma=Obst.cell(row=j+2,column=9).value
      zmi=Obst.cell(row=j+2,column=10).value
      zma=Obst.cell(row=j+2,column=11).value
      if j==0:
        Oblist=BoxBuild(xmi,xma,ymi,yma,zmi,zma)
      else:
        Oblist=np.vstack((Oblist.astype(float),BoxBuild(xmi,xma,ymi,yma,zmi,zma).astype(float)))
      # In a box all surfaces are formed of two triangles
      NTribox=2*np.ones(6)
      NtriOb=np.append(NtriOb,NTribox)
    elif Surf:
      xmi=Obst.cell(row=j+2,column=6).value
      xma=Obst.cell(row=j+2,column=7).value
      ymi=Obst.cell(row=j+2,column=8).value
      yma=Obst.cell(row=j+2,column=9).value
      zmi=Obst.cell(row=j+2,column=10).value
      zma=Obst.cell(row=j+2,column=11).value
      if abs(xmi-xma)<epsilon:
        T1 =np.array([(xmi,ymi,zmi),(xmi,yma,zmi),(xmi,yma,zma)])
        T2 =np.array([(xmi,yma,zma),(xmi,ymi,zma),(xmi,ymi,zmi)])
      elif abs(ymi-yma)<epsilon:
        T1 =np.array([(xmi,ymi,zmi),(xma,ymi,zmi),(xma,ymi,zma)])
        T2 =np.array([(xmi,ymi,zmi),(xmi,ymi,zma),(xma,ymi,zma)])
      elif abs(zmi-zma)<epsilon:
        T1 =np.array([(xmi,ymi,zmi),(xmi,yma,zmi),(xma,yma,zmi)])
        T2 =np.array([(xmi,ymi,zmi),(xma,ymi,zmi),(xma,yma,zmi)])
      else:
          raise('Surface indicated but incorrect values for bounds used')
      SurfOb=np.array([T1,T2])
      if j==0:
        Oblist=SurfOb
      else:
        Oblist=np.vstack((Oblist.astype(float),SurfOb.astype(float)))
      # In a box all surfaces are formed of two triangles
      NtriOb=np.append(NtriOb,2)
    elif Tri:
      p0=np.ones(3)
      p1=np.ones(3)
      p2=np.ones(3)
      p0[0]=Obst.cell(row=j+2,column=12).value
      p0[1]=Obst.cell(row=j+2,column=13).value
      p0[2]=Obst.cell(row=j+2,column=14).value
      p1[0]=Obst.cell(row=j+2,column=15).value
      p2[1]=Obst.cell(row=j+2,column=16).value
      p1[2]=Obst.cell(row=j+2,column=17).value
      p2[0]=Obst.cell(row=j+2,column=18).value
      p2[1]=Obst.cell(row=j+2,column=19).value
      p2[2]=Obst.cell(row=j+2,column=20).value
      Tri=np.array([p0,p1,p2]).astype(float)
      if j==1:
        Oblist=Tri
      else:
        Oblist=np.vstack((Oblist.astype(float),Tri))
      # In a box all surfaces are formed of two triangles
      NtriOb=np.append(NtriOb,1)
  for j in range(len(NtriOb)):
    NTriObSh.cell(row=j+1,column=1).value=NtriOb[j]

  #- Outer Boundary -
  # 3D co-ordinates forming a closed boundary.
  Nbox=OutB.max_row-1
  OuterBoundary=np.array([])
  for j in range(Nbox):
    ## Obstacles are all triangles in 3D.
    Box=OutB.cell(row=j+2,column=3).value
    Surf=OutB.cell(row=j+2,column=4).value
    Tri=OutB.cell(row=j+2,column=5).value
    if Box:
      xmi=OutB.cell(row=j+2,column=6).value
      xma=OutB.cell(row=j+2,column=7).value
      ymi=OutB.cell(row=j+2,column=8).value
      yma=OutB.cell(row=j+2,column=9).value
      zmi=OutB.cell(row=j+2,column=10).value
      zma=OutB.cell(row=j+2,column=11).value
      Bound=BoxBuild(xmi,xma,ymi,yma,zmi,zma)
      if j==0:
        OuterBoundary=Bound
      else:
        OuterBoundary=np.vstack((OuterBoundary,Bound))
      # In a box all surfaces are formed of two triangles
      NTribox=2*np.ones(6)
      NtriOut=np.append(NtriOut,NTribox)
    elif Surf:
      xmi=OutB.cell(row=j+2,column=6).value
      xma=OutB.cell(row=j+2,column=7).value
      ymi=OutB.cell(row=j+2,column=8).value
      yma=OutB.cell(row=j+2,column=9).value
      zmi=OutB.cell(row=j+2,column=10).value
      zma=OutB.cell(row=j+2,column=11).value
      if abs(xmi-xma)<epsilon:
        T1 =np.array([(xmi,ymi,zmi),(xmi,yma,zmi),(xmi,yma,zma)])
        T2 =np.array([(xmi,yma,zma),(xmi,ymi,zma),(xmi,ymi,zmi)])
      elif abs(ymi-yma)<epsilon:
        T1 =np.array([(xmi,ymi,zmi),(xma,ymi,zmi),(xma,ymi,zma)])
        T2 =np.array([(xmi,ymi,zmi),(xmi,ymi,zma),(xma,ymi,zma)])
      elif abs(zmi-zma)<epsilon:
        T1 =np.array([(xmi,ymi,zmi),(xmi,yma,zmi),(xma,yma,zmi)])
        T2 =np.array([(xmi,ymi,zma),(xma,ymi,zma),(xma,yma,zmi)])
      else:
          raise('Surface indicated but incorrect values for bounds used')
      SurfOb=np.array([T1,T2])
      if j==0:
        OuterBoundary=SurfOb
      else:
        OuterBoundary=np.vstack((OuterBoundary.astype(float),SurfOb.astype(float)))
      # In a box all surfaces are formed of two triangles
      NtriOut=np.append(NtriOut,2)
    elif Tri:
        p0=np.ones(3)
        p1=np.ones(3)
        p2=np.ones(3)
        p0[0]=OutB.cell(row=j+2,column=12).value
        p0[1]=OutB.cell(row=j+2,column=13).value
        p0[2]=OutB.cell(row=j+2,column=14).value
        p1[0]=OutB.cell(row=j+2,column=15).value
        p2[1]=OutB.cell(row=j+2,column=16).value
        p1[2]=OutB.cell(row=j+2,column=17).value
        p2[0]=OutB.cell(row=j+2,column=18).value
        p2[1]=OutB.cell(row=j+2,column=19).value
        p2[2]=OutB.cell(row=j+2,column=20).value
        Tri=np.array([p0,p1,p2])
        if j==0:
          OuterBoundary=Tri
        else:
          OuterBoundary=np.vstack((OuterBoundary,Tri))
        # In a box all surfaces are formed of two triangles
        NtriOut=np.append(NtriOut,1)
  for j in range(len(NtriOut)):
    NTriOutSh.cell(row=j+1,column=1).value=NtriOut[j]

  # -Router location -co-ordinate of three real numbers
  Tx=np.array([0.4,0.3,0.5])
  SimPar['D8'].value=Tx[0]
  SimPar['E8'].value=Tx[1]
  SimPar['F8'].value]=Tx[2]

  Nout =int(np.sum(NtriOut))
  Nobst=int(np.sum(NtriOb))
  Nob  =int(Nout+InnerOb*Nobst)
  ObCooStr="ObstacleCoords"
  try:
    ObstCoor=InBook[ObCooStr]
  except KeyError:
    InBook.create_sheet(ObCooStr)
    ObstCoor=InBook[ObCooStr]
  for j in range(Nob):
    if j+1<Nout:
      Obstr='OuterBoundar%03d'%j
      ObstCoor.cell(row=3*j+1,column=1).value=Obstr
      ObstCoor.cell(row=3*j+1,column=2).value=OuterBoundary[j,0,0]
      ObstCoor.cell(row=3*j+1,column=3).value=OuterBoundary[j,0,1]
      ObstCoor.cell(row=3*j+1,column=4).value=OuterBoundary[j,0,2]
      ObstCoor.cell(row=3*j+2,column=2).value=OuterBoundary[j,1,0]
      ObstCoor.cell(row=3*j+2,column=3).value=OuterBoundary[j,1,1]
      ObstCoor.cell(row=3*j+2,column=4).value=OuterBoundary[j,1,2]
      ObstCoor.cell(row=3*j+3,column=2).value=OuterBoundary[j,2,0]
      ObstCoor.cell(row=3*j+3,column=3).value=OuterBoundary[j,2,1]
      ObstCoor.cell(row=3*j+3,column=4).value=OuterBoundary[j,2,2]
    if Nout<=j+1<Nout+Nobst:
      Obstr='Obstacle%03d'%j
      ObstCoor.cell(row=3*j+1,column=1).value=Obstr
      ObstCoor.cell(row=3*j+1,column=2).value=Oblist[j-Nout,0,0]
      ObstCoor.cell(row=3*j+1,column=3).value=Oblist[j-Nout,0,1]
      ObstCoor.cell(row=3*j+1,column=4).value=Oblist[j-Nout,0,2]
      ObstCoor.cell(row=3*j+2,column=2).value=Oblist[j-Nout,1,0]
      ObstCoor.cell(row=3*j+2,column=3).value=Oblist[j-Nout,1,1]
      ObstCoor.cell(row=3*j+2,column=4).value=Oblist[j-Nout,1,2]
      ObstCoor.cell(row=3*j+3,column=2).value=Oblist[j-Nout,2,0]
      ObstCoor.cell(row=3*j+3,column=3).value=Oblist[j-Nout,2,1]
      ObstCoor.cell(row=3*j+3,column=4).value=Oblist[j-Nout,2,2]
  runplottype=SimPar.cell(row=15,column=3).value
  locatype=SimPar.cell(row=22,column=3).value
  Heatmappattern=SimPar.cell(row=16,column=3).value
  plotfit=SimPar.cell(row=24,column=3).value

  LOS=SimPar.cell(row=10,column=3).value     # LOS=1 for LOS propagation, LOS=0 for reflected propagation
  PerfRef=1
  SimPar.cell(row=11,column=3).value=PerfRef # Perfect reflection has no loss and ignores angles.
  AngChan=SimPar.cell(row=23,column=3).value # Switch for whether angles should be correction for the received point.

  # -------------------------------------------------------------------
  # CALCULATED PARAMETERS TO SAVE
  # -------------------------------------------------------------------
  # CALCULATE RELATIVE MESHWIDTH
  roomlengthscale      =l2#*abs(np.amax(OuterBoundary)-np.amin(OuterBoundary)) # SCALE WITHIN THE UNIT CO-ORDINATES.
  SimPar.cell(row=6,column=3).value=roomlengthscale
  #OuterBoundary=OuterBoundary/roomlengthscale
  #Oblist=Oblist/roomlengthscale
  h=1.0/Ns
  SimPar.cell(row=14,column=3).value=h

  if not os.path.exists('./Parameters'):
    os.makedirs('./Parameters')
  # CALCULATE ANGLE SPACING
  for j in range(0,nrays):
    xysteps       =int(ma.ceil(abs(2.0*np.pi/deltheta[j])))
    theta1        =np.linspace(0.0,2*np.pi,num=int(xysteps), endpoint=False) # Create an array of all the angles
    deltheta[j]   =theta1[1]-theta1[0]
    zsteps        =int(ma.ceil(abs(np.pi/(deltheta[j]))))
    Nra[j]=2
    Nraout=np.array([])
    start=0
    for k in range(-int(zsteps/2),int(zsteps/2)+1):
      mid=(np.cos(deltheta[j])-np.sin(k*deltheta[j])**2)/(np.cos(deltheta[j]*k)**2)
      if abs(mid)>1:
        pass
      else:
        bot=ma.acos(mid)
        xyk=int(2*np.pi/bot)
        if xyk<=1:
          break
        Nra[j]    +=xyk
        theta1    =np.linspace(0.0,2*np.pi,num=int(xyk), endpoint=False) # Create an array of all the angles
        co        =np.cos(k*deltheta[j])
        si        =np.sin(k*deltheta[j])
        updirecs  =np.c_[co*np.cos(theta1),co*np.sin(theta1),si*np.ones((xyk,1))]
        #downdirecs   =np.c_[co*np.cos(theta1),co*np.sin(theta1),-si*np.ones((xyk,1))]
        if start==0:
          coords=updirecs
          start =1
        else:
          #coords  =np.r_[coords,downdirecs]
          coords  =np.r_[updirecs,coords]
    if len(coords)<=1:
      Nraout=Nra[j+1:]
      pass
    else:
      Nraout=Nra
      directions=np.zeros((Nra[j],4))
      directions[1:-1]=np.c_[coords,np.zeros((Nra[j]-2,1))]
      directions[0] =np.array([0.0,0.0, 1.0,0.0])
      directions[-1]=np.array([0.0,0.0,-1.0,0.0])
      directionname='Parameters/Directions%03d.npy'%Nra[j]
      np.save(directionname,directions)
    Angspace.cell(row=j+2,column=2).value=Nra[j]
    Angspace.cell(row=j+2,column=1).value=deltheta[j]
    DirecStr='Directions%03d'%(Nra[j])
    try:
      DirecSh=InBook[DirecStr]
    except KeyError:
      InBook.create_sheet(DirecStr)
      DirecSh=InBook[DirecStr]
    for i in range(int(Nra[j])):
      DirecSh.cell(row=i+1,column=1).value=directions[i,0]
      DirecSh.cell(row=i+1,column=2).value=directions[i,1]
      DirecSh.cell(row=i+1,column=3).value=directions[i,2]
    np.savetxt('Parameters/'+DirecStr+'.csv', directions, delimiter=',', fmt='%f')

  # COMBINE THE RAY-LAUNCHER PARAMETERS INTO ONE ARRAY
  RTPar=np.array([Nre,h,roomlengthscale,split])

  #print('Origin of raytracer ', Tx)

  # --------------------------------------------------------------------
  # SAVE THE PARAMETERS IN A FOLDER TITLED `Parameters`
  # --------------------------------------------------------------------
  np.save('Parameters/ResOn.npy',ResOn)
  np.save('Parameters/logon.npy',logon)
  np.save('Parameters/MaxInter.npy',MaxInter)
  np.save('Parameters/testnum.npy',testnum)
  np.save('Parameters/roomnumstat.npy',roomnumstat)
  np.save('Parameters/timetest.npy',timetest)
  np.save('Parameters/Raytracing.npy',RTPar)
  np.save('Parameters/Nra.npy',Nraout)
  np.save('Parameters/delangle.npy',deltheta)
  if InnerOb:
    Oblist=np.concatenate((OuterBoundary,Oblist))
  else:
    Oblist=OuterBoundary
  np.save('Parameters/Obstacles.npy',Oblist)
  np.save('Parameters/NtriOb.npy',NtriOb)
  np.save('Parameters/InnerOb.npy',InnerOb)
  np.save('Parameters/OuterBoundary.npy',OuterBoundary)
  np.save('Parameters/NtriOut.npy',NtriOut)
  np.save('Parameters/Nob.npy',Nob)
  np.save('Parameters/Ns.npy',Ns)
  np.save('Parameters/Nrs.npy',Nrs)
  np.save('Parameters/Nob.npy',Nob)
  np.save('Parameters/Origin.npy',Tx)
  np.save('Parameters/LOS.npy',LOS)
  np.save('Parameters/PerfRef.npy',PerfRef)
  np.save('Parameters/AngChan.npy',AngChan)

  nsur=len(NtriOb)

  if LOS:
    LOSstr='LOS'
  elif PerfRef:
    if Nre>2 and Nrs>1:
      if Nrs<nsur:
       LOSstr=nw.num2words(Nrs)+'PerfRef'
      else:
       LOSstr='MultiPerfRef'
    else:
      LOSstr='SinglePerfRef'
  else:
    if Nre>2 and Nrs>1:
      if Nrs<nsur:
       LOSstr=nw.num2words(Nrs)+'Ref'
      else:
       LOSstr='MultiRef'
    else:
      LOSstr='SingleRef'
  if InnerOb:
    boxstr='Box'
  else:
    boxstr='NoBox'
  if abs(Tx[0]-0.5)<epsilon and abs(Tx[1]-0.5)<epsilon and abs(Tx[2]-0.5)<epsilon:
    loca='Centre'
  else:
    loca='OffCentre'
  SimPar.cell(row=22,column=3).value=loca
  runplottype=LOSstr+boxstr+loca
  SimPar.cell(row=15,column=3).value=runplottype
  text_file = open('Parameters/runplottype.txt', 'w')
  n = text_file.write(runplottype)
  text_file.close()
  text_file = open('Parameters/locatype.txt', 'w')
  n = text_file.write(locatype)
  text_file.close()
  text_file = open('Parameters/Heatmapstyle.txt', 'w')
  n = text_file.write(Heatmappattern)
  text_file.close()
  text_file = open('Parameters/PlotFit.txt', 'w')
  n = text_file.write(plotfit)
  text_file.close()
  #print('------------------------------------------------')
  #print('Geometrical parameters saved')
  #print('------------------------------------------------')
  InBook.save(filename=SN)
  return 0

def ObstacleCoefficients(SN,index=0):
  ''' Input the paramters for obstacles and the antenna. To ensure \
  arrays are of the right length for compatibility for the \
  ray-launcher retrieve the ray-launching parameters in \
  :py:func:`DeclareParameters()`

  Load:

  * 'Obstacles.npy'     -Co-ordinates of obstacles in the room
  * 'OuterBoundary.npy' - Co-ordinates of the walls of the room
  * 'Raytracing.npy'    -[Nra (number of rays), Nre (number of reflections), \
  h (relative meshwidth)]

  Calculate:

  * Nob=len([Obstacles,OuterBoundary])

  Input:

  * `Freespace` -[mu0 (permeability of air), \
  eps0 (permittivity of air),Z0 (characteristic impedance of air), \
  c (speed of light)]
  * `frequency` - :math:`\\omega` angular frequency of the wave out \
  the antenna.
  * `mur`       - :math:`\\mu_r` The relative permeability for all obstacles. \
  This should be an array with the same number of terms as the number \
  of obstacles Nob.
  * `epsr`     - :math:`\\epsilon_r` The relative permittivity for each obstacle. \
  This should be an array with the same number of terms as the number \
  of obstacles Nob.
  * `sigma`     - :math:`\\sigma` The electrical conductivity of the obstacles. \
  This should be an array with the same number of terms as the number \
  of obstacles.
  * `Gt`        - The gains of the antenna. The should be an array with \
  the same number of terms as the number of rays Nra.

  Calculate:

  * `eps0`   - :math:`\\epsilon_0=\\frac{1}{\\mu_0 c^2}`  permittivity of \
  freespace.
  * `Z0`     - :math:`Z_0=\\sqrt{\\frac{\\mu_0}{\\epsilon_0}}` characteristic \
  impedance of freespace.
  * `refindex` - The refreactive index \
  :math:`n=\\sqrt{\\mu_r\\epsilon_r}`
  * `Znobrat`- The relative impedance of the obstacles given by,
    :math:`\\hat{Z}_{Nob}=\\frac{Z_{Nob}}{Z_0}`. The impedance of each \
    obstacle :math:`Z_{Nob}` is given by \
    :math:`Z_{Nob}=\\sqrt{\\frac{i\\omega\\mu_0\\mu_r}{\\sigma+i\\epsilon_0\\epsilon_r}}`.


  The Znobrat and refindex terms are then reformatted so that they \
  repeat Nre times with an extra term. The extra term corresponds to \
  the line of sight path. This makes them the same length as a column \
  in a matrix in a :py:class:`DictionarySparseMatrix.DS`. \
  Each term corresponds to a possible obstacle reflection combination.

  The Gains matrix is also reformated to that it repeats (Nre+1) times. \
  This corresponds to every possible ray reflection number combination \
  This makes them the same length as a row in a matrix in a \
  :py:class:`DictionarySparseMatrix.DS`. \
  Each term corresponds to a possible obstacle reflection combination.

  Save:
  * `frequency.npy`- The angular frequency :math:`\\omega`.
  * `refindex.npy` - The refractive index of the obstacles.
  * `Znobrat.npy`  - The relative characteristic impedance.
  * `TxGains.npy`  - The gains of the antenna.
  * `Freespace.npy`- The freespace parameters.

  .. code::

     Freespace=np.array([mu0,eps0,Z0,c])

  :return: 0 if successfully completed.

  '''
  #print('Saving the physical parameters for obstacles and antenna')

  # -------------------------------------------------------------------
  # RETRIEVE RAY LAUNCHER PARAMETERS FOR ARRAY LENGTHS-----------------
  # -------------------------------------------------------------------
  InBook     =wb.load_workbook(filename=SN,data_only=True)
  AirMat     ='AirParameters'
  ObstMat    ='ObstacleMaterial'
  Simparstr  ='SimulationParameters'
  Obststr    ='Obstacles'
  OutBstr    ='OuterBoundary'
  Obst       =InBook[Obststr]
  OutB       =InBook[OutBstr]
  Air        =InBook[AirMat]
  ObstPar    =InBook[ObstMat]
  SimPar     =InBook[Simparstr]
  if not os.path.exists('./Parameters/'):
    os.makedirs('./Parameters/')
  RTPar         =np.load('Parameters/Raytracing.npy')
  Nre,h,L,split =RTPar
  Oblist        =np.load('Parameters/Obstacles.npy')
  Nra           =np.load('Parameters/Nra.npy')
  if isinstance(Nra, (float,int,np.int32,np.int64, np.complex128 )):
      nra=np.array([Nra])
  else:
      nra=len(Nra)
  Nre=int(RTPar[0])                           # Number of reflections
  Nob=np.load('Parameters/Nob.npy')           # The Number of obstacle.
  Nrs =SimPar.cell(row=13,column=3).value

  # -------------------------------------------------------------------
  # INPUT PARAMETERS FOR POWER CALCULATIONS----------------------------
  # -------------------------------------------------------------------

  # PHYSICAL CONSTANTS Do not change for obstacles for frequency
  mu0=complex(Air.cell(row=2,column=1).value) #4*np.pi*1E-6
  c  =float(Air.cell(row=2,column=3).value) #2.99792458E+8


  # ANTENNA PARAMETERS
  #-----------------------------------------------------------------------
  # Gains of the rays
  for j in range(0,nra):
    Gt=np.ones((Nra[j],1),dtype=np.complex128)
    gainname='Parameters/Tx%03dGains%03d.npy'%(Nra[j],index)
    np.save(gainname, Gt)
  frequency     =float(Air.cell(row=2,column=4).value)              # 2.79 GHz #FIXME make this a table and choose a frequency option
  khat          =frequency*L/c                 # Non-dimensional wave number
  Air.cell(row=2,column=5).value=khat
  lam           =(2*np.pi)/khat                # Non-dimensional wavelength
  Air.cell(row=2,column=6).value=lam
  Antpar        =np.array([khat,lam,L])
  Pol           =np.array([Air.cell(row=2,column=7).value,Air.cell(row=3,column=7).value])
  # CALCULATE FREESPACE PARAMETERS

  eps0=1/(mu0*c**2)  #8.854187817E-12
  Air.cell(row=2,column=2).value=str(eps0)
  Z0=(mu0/eps0)**0.5 #120*np.pi Characteristic impedance of free space.
  Air.cell(row=2,column=8).value=str(Z0)

  # STORE FREESPACE PARAMETERS IN ONE ARRAY
  Freespace=np.array([mu0,eps0,Z0,c])
  # OBSTACLE CONTSTANTS
  #----------------------------------------------------------------------
  # Relative Constants for the obstacles
  NtriOut=np.load('Parameters/NtriOut.npy')
  NtriOb =np.load('Parameters/NtriOb.npy')

  InOb=SimPar.cell(row=7,column=3).value
  NOut=OutB.max_row-1
  Nobst=Obst.max_row-1
  Nsur=Nobst*InOb+NOut
  mur   =np.array([complex(ObstPar.cell(row=2,column=2).value)])
  epsr  =np.array([complex(ObstPar.cell(row=2,column=3).value)])
  sigma =np.array([complex(ObstPar.cell(row=2,column=4).value)])
  top=frequency*mu0*mur[-1]*1j
  bottom=sigma[-1]+eps0*frequency*epsr[-1]*1j
  if bottom==0:
    Znob=0
  else:
    Znob =np.sqrt(top/bottom)                    # Wave impedance of the obstacles
  Znobrat=np.array([Znob/Z0])
  ObstPar.cell(row=2,column=5).value=str(Znob/Z0)
  PerfRef=SimPar.cell(row=11,column=3).value
  if PerfRef and Znob!=0:
    refindec=np.array([1+0j])
  else:
    refindex=np.array([mur[-1]*epsr[-1]])
  ObstPar.cell(row=2,column=6).value=str(refindex[-1])
  nre=0
  for j in range(1,Nsur):
    box=int(OutB.cell(row=j+2,column=3).value)
    surf=int(OutB.cell(row=j+2,column=4).value)
    tri=int(OutB.cell(row=j+2,column=5).value)
    if j+1<=NOut:
      for i in range(1,6*box+2*surf+tri):
        if j+2==4:
          murj =complex(ObstPar.cell(row=j+3,column=2).value)
          epsrj=complex(ObstPar.cell(row=j+3,column=3).value)
          sigj =complex(ObstPar.cell(row=j+3,column=4).value)
        murj =complex(ObstPar.cell(row=j+2,column=2).value)
        epsrj=complex(ObstPar.cell(row=j+2,column=3).value)
        sigj =complex(ObstPar.cell(row=j+2,column=4).value)
        mur   =np.append(mur  ,murj)
        epsr  =np.append(epsr ,epsrj)
        sigma =np.append(sigma,sigj)
        top=frequency*mu0*murj*1j
        bottom=sigj+eps0*frequency*epsrj*1j
        if bottom==0:
          Znob=0
        else:
          nre+=1
          Znob   =np.sqrt(top/bottom)                    # Wave impedance of the obstacles
        Znobrat=np.append(Znobrat,Znob/Z0)
        ObstPar.cell(row=j+2,column=5).value=str(Znob/Z0)
        if nre > Nrs+1:
          refindex=np.append(refindex,0+0j)
          ObstPar.cell(row=j+2,column=6).value=str(0+0j)
        else:
          if PerfRef and Znob!=0:
            refindex=np.append(refindex,1)
            ObstPar.cell(row=j+2,column=6).value=str(1+0j)
          else:
            refindex=np.append(refindex,murj*epsrj)
            ObstPar.cell(row=j+2,column=6).value=str(murj*epsrj)
    if NOut+Nobst+1>j+1>NOut:
      for i in range(1,box*6+surf*2+tri):
        murj =complex(ObstPar.cell(row=j+2,column=2).value)
        epsrj=complex(ObstPar.cell(row=j+2,column=3).value)
        sigj =complex(ObstPar.cell(row=j+2,column=4).value)
        mur   =np.append(mur  ,murj)
        epsr  =np.append(epsr ,epsrj)
        sigma =np.append(sigma,sigj)
        top=frequency*mu0*murj*1j
        bottom=sigj+eps0*frequency*epsrj*1j
        if bottom==0:
          Znob=0
        else:
          nre+=1
          Znob   =np.sqrt(top/bottom)
        Znobrat=np.append(Znobrat,Znob/Z0)
        ObstPar.cell(row=j+2,column=5).value=str(Znob/Z0)
        if nre>Nrs+1:
          refindex=np.append(refindex,0+0j)
          ObstPar.cell(row=j+2,column=6).value=str(0+0j)
        else:
          if PerfRef and Znob!=0:
            refindex=np.append(refindex,1)
            ObstPar.cell(row=j+2,column=6).value=str(1+0j)
          else:
            refindex=np.append(refindex,murj*epsrj)
            ObstPar.cell(row=j+2,column=6).value=str(murj*epsrj)

  # --------------------------------------------------------------------
  # SAVE THE PARAMETERS
  # --------------------------------------------------------------------
  Obstr=''
  if Nrs<Nsur:
    obnumbers=np.zeros((Nrs,1))
    k=0
    for ob, refin in enumerate(refindex):
      if abs(refin)>epsilon:
        obnumbers[k]=ob
        k+=1
        Obstr=Obstr+'Ob%02d'%ob
  text_file = open('Parameters/Obstr.txt', 'w')
  n = text_file.write(Obstr)
  text_file.close()
  SimPar.cell(row=26,column=3).value=Obstr
  np.save('Parameters/Freespace%03d.npy'%index,Freespace)
  np.save('Parameters/frequency%03d.npy'%index,frequency)
  np.save('Parameters/lam%03d.npy'%index,lam)
  np.save('Parameters/khat%03d.npy'%index,khat)
  np.save('Parameters/Antpar%03d.npy'%index,Antpar)
  np.save('Parameters/Znobrat%03d.npy'%index,Znobrat)
  np.save('Parameters/refindex%03d.npy'%index,refindex)
  np.save('Parameters/Pol%03d.npy'%index,Pol)
  #print('------------------------------------------------')
  #print('Material parameters saved')
  #print('------------------------------------------------')
  InBook.save(filename=SN)
  return 0

def BoxBuild(xmi,xma,ymi,yma,zmi,zma):
  ''' Input the inimum and maximum x,y, and z co-ordinates which will form a Box.
  :param xmi: The minimum x co-ordinate.
  :param xma: The maximum x co-ordinate.
  :param ymi:The minimum y co-ordinate.
  :param yma: The maximum y co-ordinate.
  :param zmi: The minimum z co-ordinate.
  :param zma: The maximum z co-ordinate.

  .. code::

       Box=[T0,T1,...T12]
       TJ=[p0J,p1J,p2J]
       p0J=[x0J,y0J,z0J]
       p1J=[x1J,y1J,z1J]
       p2J=[x2J,y2J,x2J]

  :rtype: 12 x 3 x 3 numpy array.
  :returns: Box
  '''
  # The faces in the y=ymi plane
  triangle1 =np.array([(xmi,ymi,zmi),(xma,ymi,zmi),(xmi,ymi,zma)])
  triangle2 =np.array([(xmi,ymi,zma),(xma,ymi,zma),(xma,ymi,zmi)])
  # The faces in the x=xmi plane
  triangle3 =np.array([(xmi,ymi,zmi),(xmi,yma,zmi),(xmi,ymi,zma)])
  triangle4 =np.array([(xmi,ymi,zma),(xmi,yma,zmi),(xmi,yma,zma)])
  # The faces in the z=zmi plane
  triangle5 =np.array([(xmi,ymi,zmi),(xma,yma,zmi),(xmi,yma,zmi)])
  triangle6 =np.array([(xmi,ymi,zmi),(xma,yma,zmi),(xma,ymi,zmi)])
  # The faces in the z=zma plane
  triangle7 =np.array([(xmi,ymi,zma),(xma,yma,zma),(xma,ymi,zma)])
  triangle8 =np.array([(xmi,ymi,zma),(xma,yma,zma),(xmi,yma,zma)])
  # The faces in the x=xma plane
  triangle9 =np.array([(xma,ymi,zma),(xma,yma,zma),(xma,yma,zmi)])
  triangle10=np.array([(xma,ymi,zma),(xma,ymi,zmi),(xma,yma,zmi)])
  # The faces in the y=yma plane
  triangle11=np.array([(xma,yma,zmi),(xma,yma,zma),(xmi,yma,zma)])
  triangle12=np.array([(xma,yma,zmi),(xmi,yma,zmi),(xmi,yma,zma)])
  # Put the triangular faces into an array called Box.
  Box=np.array([triangle1,triangle2,triangle3,triangle4,triangle5,
  triangle6,triangle7,triangle8,triangle9,triangle10,triangle11,
  triangle12])
  return Box

def initialload(index=0):
  mu0,eps0,Z0,c=np.load('Parameters/Freespace%03d.npy'%index)
  freq     =np.load('Parameters/frequency%03d.npy'%index)
  lam=np.load('Parameters/lam%03d.npy'%index)
  khat=np.load('Parameters/khat%03d.npy'%index)
  _,_,L=np.load('Parameters/Antpar%03d.npy'%index)
  Znobrat=np.load('Parameters/Znobrat%03d.npy'%index)
  refindex=np.load('Parameters/refindex%03d.npy'%index)
  Pol=np.load('Parameters/Pol%03d.npy'%index)

  ResOn                       =np.load('Parameters/ResOn.npy')
  logon                       =np.load('Parameters/logon.npy')
  MaxInter                    =np.load('Parameters/MaxInter.npy')
  testnum                     =np.load('Parameters/testnum.npy')
  roomnumstat                 =np.load('Parameters/roomnumstat.npy')
  timetest                    =np.load('Parameters/timetest.npy')
  Nre,h,roomlengthscale,split =np.load('Parameters/Raytracing.npy')
  Nra                         =np.load('Parameters/Nra.npy')
  deltheta                    =np.load('Parameters/delangle.npy')
  Oblist                      =np.load('Parameters/Obstacles.npy')
  NtriOb                      =np.load('Parameters/NtriOb.npy')
  InnerOb                     =np.load('Parameters/InnerOb.npy')
  OutBoundary                 =np.load('Parameters/OuterBoundary.npy')
  NtriOut                     =np.load('Parameters/NtriOut.npy')
  Nob                         =np.load('Parameters/Nob.npy')
  Ns                          =np.load('Parameters/Ns.npy')
  Nrs                         =np.load('Parameters/Nrs.npy')
  Tx                          =np.load('Parameters/Origin.npy')
  LOS                         =np.load('Parameters/LOS.npy')
  PerfRef                     =np.load('Parameters/PerfRef.npy')
  AngChan                     =np.load('Parameters/AngChan.npy')
  myfile = open('Parameters/runplottype.txt', 'rt') # open lorem.txt for reading text
  plottype= myfile.read()         # read the entire file into a string
  myfile.close()
  logname='Parameters'+plottype+'.log'
  logging.basicConfig(filename=logname,filemode='w',format="[%(asctime)s ]%(message)s",
    level=logging.INFO,
    datefmt='%Y-%m-%d %H:%M:%S')
  logging.info(sys.version)
  if ResOn:
    msg='Residual to be calculated'
  else:
    msg='Residual calculation turned off'
  print(msg)
  logging.info(msg)
  if logon:
    msg='logging to be done'
  else:
    msg='logging turned off for speed.'
  print(msg)
  logging.info(msg)
  if InnerOb:
    msg='Interior obstacles included'
  else:
    msg='Only the bounding walls included'
  print(msg)
  logging.info(msg)
  if  LOS:
    msg='Only line of sight paths included'
  else:
    msg='LOS and refelected paths included'
  print(msg)
  logging.info(msg)
  if PerfRef:
    msg='Reflections are modelled as perfect with no loss'
  else:
    msg='loss is accounted for at reflection with angles used'
  print(msg)
  logging.info(msg)
  if AngChan:
    msg='Angles of reflection are corrected when the receiver is moved through cones or centre point'
  else:
    msg='Reflection angles for all points in the cone are determined by the ray'
  print(msg)
  logging.info(msg)

  print('Permittivity in air',mu0)
  print('Permeability in air',eps0)
  print('Impedance of air',Z0)
  print('Speed of light',c)
  print('Frequency of EM wave',freq)
  print('Wavelength',lam)
  print('Non-dimensional wave number',khat)
  print('Environment lengthscale',L)
  print('Ratio of impedance of obstacles over impedance of air',Znobrat)
  print('Refractive index of the obstacles',refindex)
  print('Polarisation',Pol)
  print('Maximum number of intersections for any ray ',MaxInter)
  print('Number of runs of the algorithm for roomtypes ',testnum)
  print('Starting number of different rooms ',roomnumstat)
  print('Number of times to repeat the algorithm for testing times ',timetest)
  print('Number of reflections ',Nre)
  print('Mesh width ',h)
  print('Room lengthscale ',L)
  print('Steps through each voxel ',split)
  print('Number of rays ',Nra)
  print('Angle spacing ',deltheta)
  print('All obstacles in the environment ',Oblist)
  print('Number of triangles forming each obstacle ',  NtriOb)
  print('The triangles forming the boundary of the environment ',OutBoundary)
  print(' The number of triangles per boundary surface ',  NtriOut)
  print('The total number of triangles forming obstacles in the environment ', Nob)
  print('The number of steps to discretise the longest axis into ', Ns)
  print('The number of reflective surface in the enironment (all others are totally absorbing) ',Nrs)
  print('Position of the transmitter ',Tx)
  logging.info('Permittivity in air %f+i%f'%(mu0.real,mu0.imag))
  logging.info('Permeability in air %f+i%f'%(eps0.real,eps0.imag))
  logging.info('Impedance of air %f+i%f'%(Z0.real,Z0.imag))
  logging.info('Speed of light %f+i%f'%(c.real,c.imag))
  logging.info('Frequency of EM wave %f'%freq)
  logging.info('Wavelength %f'%lam)
  logging.info('Non-dimensional wave number %f'%khat)
  logging.info('Environment lengthscale %f'%L)
  logging.info('Ratio of impedance of obstacles over impedance of air '+str(Znobrat))
  logging.info('Refractive index of the obstacles '+str(refindex))
  logging.info('Polarisation (%d,%d)'%(Pol[0],Pol[1]))
  logging.info('Maximum number of intersections for any ray %d'%MaxInter)
  logging.info('Number of runs of the algorithm for roomtypes %d'%testnum)
  logging.info('Starting number of different rooms %d'%roomnumstat)
  logging.info('Number of times to repeat the algorithm for testing times %d'%timetest)
  logging.info('Number of reflections %d'%Nre)
  logging.info('Mesh width %f'%h)
  logging.info('Room lengthscale %f'%L)
  logging.info('Steps through each voxel %d'%split)
  logging.info('Number of rays '+str(Nra))
  logging.info('Angle spacing '+str(deltheta))
  logging.info('All obstacles in the environment '+str(Oblist))
  logging.info('Number of triangles forming each obstacle '+str(NtriOb))
  logging.info('The triangles forming the boundary of the environment '+str(OutBoundary))
  logging.info(' The number of triangles per boundary surface '+str(NtriOut))
  logging.info('The total number of triangles forming obstacles in the environment %d'%Nob)
  logging.info('The number of steps to discretise the longest axis into %d'%Ns)
  logging.info('The number of reflective surface in the enironment (all others are totally absorbing) %d'%Nrs)
  logging.info('Position of the transmitter (%f,%f,%f)'%(Tx[0],Tx[1],Tx[2]))
  return

if __name__=='__main__':
  np.set_printoptions(precision=3)
  print('Running  on python version')
  print(sys.version)
  Sheetname='InputSheet.xlsx'
  out=DeclareParameters(Sheetname)
  out=ObstacleCoefficients(Sheetname)
  initialload()

  exit()



